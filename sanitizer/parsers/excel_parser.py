"""
Excel 文件解析器
支持 .xlsx, .xls, .csv, .tsv 格式
"""

import csv
import io
import logging
import math
import shutil
import subprocess
import tempfile
from datetime import datetime, date, time
from pathlib import Path

from sanitizer.engine import SanitizeEngine
from sanitizer.parsers.base import BaseParser

logger = logging.getLogger(__name__)


def _patch_openpyxl() -> None:
    """
    修补 openpyxl 对 CalculatedItem.formula 的严格类型校验。

    问题根源：
      部分财务系统（SAP、用友等）导出的 xlsx 中，数据透视表的
      <calculatedItem> 元素没有 formula 属性，openpyxl 解析时
      将其传入 __init__ 为 None，赋值给 String 描述符时报 TypeError。

    需要修补两个不同的 CalculatedItem 类：
      - openpyxl.pivot.cache.CalculatedItem  ← pivotCacheDefinition 解析时触发
      - openpyxl.pivot.table.CalculatedItem  ← pivotTableDefinition 解析时触发
    两者都有同样的 formula: String 描述符，都会在 formula=None 时报错。

    修补方式：
      替换 CalculatedItem.__init__，在调用原始 __init__ 前把
      formula=None 改为 formula=""，完全绕过描述符校验。
      不影响 openpyxl 其他任何类或行为。

    时机：
      模块导入时执行一次，对整个进程生效。
    """
    # 需要修补的两个类的导入路径
    patch_targets = [
        "openpyxl.pivot.cache",   # pivotCacheDefinition — 实际报错的来源
        "openpyxl.pivot.table",   # pivotTableDefinition — 也可能触发同样问题
    ]

    patched_count = 0
    for module_path in patch_targets:
        try:
            import importlib
            mod = importlib.import_module(module_path)
            cls = getattr(mod, "CalculatedItem", None)
            if cls is None:
                continue
            if getattr(cls, "_formula_patched", False):
                patched_count += 1
                continue

            _orig_init = cls.__init__

            def _make_patched(orig):
                def _patched_init(self, formula=None, **kwargs):
                    orig(self, formula=formula if formula is not None else "", **kwargs)
                return _patched_init

            cls.__init__ = _make_patched(_orig_init)
            cls._formula_patched = True
            patched_count += 1
            logger.debug(f"已修补 {module_path}.CalculatedItem.__init__")

        except Exception as e:
            logger.debug(f"{module_path}.CalculatedItem 补丁未能应用: {e}")

    if patched_count > 0:
        logger.debug(f"openpyxl CalculatedItem 补丁: 共修补 {patched_count} 个类")
    else:
        logger.debug("openpyxl CalculatedItem 补丁: 未找到需要修补的类")


# 模块加载时立即执行
_patch_openpyxl()


class ExcelParser(BaseParser):
    """
    Excel/CSV 文件解析器

    - .xlsx: 使用 openpyxl 读写，保留样式、合并单元格、sheet 结构
    - .xls: 使用 xlrd 读取，openpyxl 写出为 .xlsx
    - .csv/.tsv: 使用 csv 模块处理，chardet 检测编码
    """

    EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv"}

    def can_handle(self, file_path: str) -> bool:
        return Path(file_path).suffix.lower() in self.EXTENSIONS

    def sanitize(self, input_path: str, output_path: str, engine: SanitizeEngine) -> dict:
        ext = Path(input_path).suffix.lower()
        engine.set_current_file(input_path)

        if ext == ".xlsx":
            return self._process_xlsx(input_path, output_path, engine)
        elif ext == ".xls":
            return self._process_xls(input_path, output_path, engine)
        elif ext in (".csv", ".tsv"):
            return self._process_csv(input_path, output_path, engine, ext)
        else:
            return {"entities_found": 0, "entities_replaced": 0}

    def _process_xlsx(self, input_path: str, output_path: str, engine: SanitizeEngine) -> dict:
        """处理 .xlsx 文件"""
        import openpyxl

        stats_before = dict(engine.stats)

        # 加载策略（依次降级）：
        #   1. 正常模式 — 保留全部样式、透视表
        #      （_patch_openpyxl 已修复 CalculatedItem.formula=None 问题，
        #        大多数之前报错的文件现在能在此步成功）
        #   2. data_only 模式 — 跳过公式校验
        #   3. LibreOffice 重新导出 — 修复更深层 XML 损坏，再正常加载
        #   4. read_only 流式 — 最终兜底（样式/透视表丢失，但数据正确）
        fname = Path(input_path).name
        wb = None
        read_only_mode = False
        lo_sheet_dims = None   # LibreOffice 转换前的行列数，用于完整性校验

        # ── 步骤 1 & 2：正常模式 / data_only ──────────────────────────────
        for kwargs, desc in [
            ({}, "正常模式"),
            ({"data_only": True}, "data_only 模式"),
        ]:
            try:
                wb = openpyxl.load_workbook(input_path, **kwargs)
                if desc != "正常模式":
                    logger.warning(f"[{fname}] 正常模式失败，已降级使用 {desc}")
                break
            except Exception as e:
                logger.warning(f"[{fname}] {desc} 加载失败 → {type(e).__name__}: {e}")
                continue

        # ── 步骤 3：LibreOffice 重新导出 ───────────────────────────────────
        if wb is None:
            if self._find_libreoffice():
                logger.warning(f"[{fname}] 尝试 LibreOffice 重新导出修复...")
                lo_buf = self._try_libreoffice_convert(input_path)
                if lo_buf is not None:
                    # 记录原始行列数用于完整性校验
                    try:
                        orig_wb = openpyxl.load_workbook(
                            input_path, read_only=True, data_only=True
                        )
                        lo_sheet_dims = {
                            s: (orig_wb[s].max_row, orig_wb[s].max_column)
                            for s in orig_wb.sheetnames
                        }
                        orig_wb.close()
                    except Exception:
                        pass

                    try:
                        wb = openpyxl.load_workbook(lo_buf)
                        logger.warning(f"[{fname}] LibreOffice 重新导出成功，已正常加载")
                    except Exception as e:
                        logger.warning(
                            f"[{fname}] LibreOffice 导出后仍无法加载 → "
                            f"{type(e).__name__}: {e}"
                        )
                else:
                    logger.warning(f"[{fname}] LibreOffice 转换失败，继续降级")
            else:
                logger.warning(
                    f"[{fname}] 未检测到 LibreOffice（可安装以提升兼容性），跳过此步骤\n"
                    f"          安装方式：brew install --cask libreoffice"
                )

        # ── 步骤 4：calamine 高速读取 ──────────────────────────────────────
        # python-calamine 是 Rust 引擎，读取速度比 openpyxl read_only 快 20-30 倍
        # 仅提取数据值，样式/透视表不保留（与 read_only 相同）
        if wb is None:
            calamine_result = self._try_calamine(input_path, output_path, engine, stats_before)
            if calamine_result is not None:
                return calamine_result

        # ── 步骤 5：read_only 流式兜底（最慢，最后手段）─────────────────────
        if wb is None:
            try:
                wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
                read_only_mode = True
                logger.warning(
                    f"[{fname}] 已降级为只读流式模式（⚠️ 输出文件将丢失样式和透视表）\n"
                    f"          建议：用 Excel 打开原文件另存为新 .xlsx 后再处理"
                )
            except Exception as e:
                logger.error(
                    f"[{fname}] read_only 模式也失败 → {type(e).__name__}: {e}"
                )

        # ── 所有模式均失败 ─────────────────────────────────────────────────
        if wb is None:
            logger.error(
                f"[{fname}] ❌ 所有加载模式均失败，跳过此文件\n"
                f"          可能原因：文件严重损坏 / 加密 / 格式不受支持\n"
                f"          建议：用 Excel 打开修复后重试"
            )
            return {"entities_found": 0, "entities_replaced": 0}

        if read_only_mode:
            # read_only 模式下 cell 不可修改，需复制到新 workbook 再处理
            new_wb = openpyxl.Workbook()
            if new_wb.sheetnames:
                new_wb.remove(new_wb.active)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                new_ws = new_wb.create_sheet(title=sheet_name)
                scale_seed = f"{Path(input_path).name}::{sheet_name}"
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if val is None:
                            continue
                        if isinstance(val, (datetime, date, time)):
                            new_ws.cell(row=cell.row, column=cell.column, value=val)
                        elif isinstance(val, str) and val.startswith("="):
                            new_ws.cell(row=cell.row, column=cell.column, value=val)
                        elif isinstance(val, bool):
                            new_ws.cell(row=cell.row, column=cell.column, value=val)
                        elif isinstance(val, str) and val.strip():
                            new_ws.cell(row=cell.row, column=cell.column,
                                        value=engine.process_text(val))
                        elif isinstance(val, (int, float)):
                            fval = float(val)
                            if not (2000 <= fval <= 2099):
                                val = engine.scale_number(val, scale_seed)
                            new_ws.cell(row=cell.row, column=cell.column, value=val)
                        else:
                            new_ws.cell(row=cell.row, column=cell.column, value=val)
            wb.close()
            wb = new_wb

        else:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                scale_seed = f"{Path(input_path).name}::{sheet_name}"

                # 重置维度，防止某些文件 max_row 虚高（如设为 1048576）导致迭代百万空行
                try:
                    ws.reset_dimensions()
                except Exception:
                    pass

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue

                        if isinstance(cell.value, (datetime, date, time)):
                            continue
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            continue
                        if isinstance(cell.value, bool):
                            continue

                        if isinstance(cell.value, str) and cell.value.strip():
                            cell.value = engine.process_text(cell.value)
                        elif isinstance(cell.value, (int, float)):
                            fmt = cell.number_format or ""
                            if self._is_date_format(fmt):
                                continue
                            val = float(cell.value)
                            if 2000 <= val <= 2099:
                                continue
                            cell.value = engine.scale_number(cell.value, scale_seed)

        # LibreOffice 转换完整性校验：对比行列数，差异超过 5% 则警告
        if lo_sheet_dims and not read_only_mode:
            for sheet_name, (orig_rows, orig_cols) in lo_sheet_dims.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"[完整性] Sheet 丢失: {fname} → {sheet_name}")
                    continue
                ws = wb[sheet_name]
                new_rows = ws.max_row or 0
                new_cols = ws.max_column or 0
                if orig_rows and new_rows and abs(new_rows - orig_rows) / max(orig_rows, 1) > 0.05:
                    logger.warning(
                        f"[完整性] {fname} Sheet '{sheet_name}' 行数变化较大: "
                        f"{orig_rows} → {new_rows}，请人工核查"
                    )

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        wb.close()

        return self._calc_stats(engine.stats, stats_before)

    @staticmethod
    def _is_date_format(fmt: str) -> bool:
        """
        判断 Excel number_format 是否为日期/时间格式

        常见日期格式包含: y, m, d, h, s 等日期/时间占位符
        但 "0.00", "#,##0" 等数值格式也包含这些字母
        """
        if not fmt or fmt == "General":
            return False

        # 常见的日期格式关键词
        date_indicators = [
            "yy", "yyyy", "mm", "dd",
            "h:mm", "hh:mm", "ss",
            "AM/PM", "am/pm",
            "m/d", "d/m", "y/m",
            "mmm", "mmmm",  # 月份名称
            "ddd", "dddd",  # 星期名称
        ]
        fmt_lower = fmt.lower()
        return any(ind.lower() in fmt_lower for ind in date_indicators)

    # ── LibreOffice 辅助方法 ───────────────────────────────────────────────

    @staticmethod
    def _find_libreoffice() -> str | None:
        """
        查找系统中的 LibreOffice 可执行文件路径

        搜索顺序：
        1. PATH 中的 libreoffice / soffice
        2. macOS 默认安装路径（/Applications/LibreOffice.app）
        3. Windows 默认安装路径（Program Files）
        返回可执行文件路径，未找到则返回 None
        """
        # PATH 中查找
        for cmd in ("libreoffice", "soffice"):
            if shutil.which(cmd):
                return cmd

        # macOS 固定路径
        mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        if Path(mac_path).exists():
            return mac_path

        # Windows 固定路径
        for win_path in (
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ):
            if Path(win_path).exists():
                return win_path

        return None

    @staticmethod
    def _try_libreoffice_convert(input_path: str) -> io.BytesIO | None:
        """
        用 LibreOffice 将 xlsx 文件重新导出，修复内部 XML 结构问题

        原理：LibreOffice 的解析器兼容性更强，能读取 openpyxl 严格校验会报错的文件，
        重新导出后生成符合标准的 xlsx，openpyxl 可正常加载。

        返回修复后文件的 BytesIO，失败（未安装/转换出错）则返回 None
        """
        lo = ExcelParser._find_libreoffice()
        if lo is None:
            return None

        fname = Path(input_path).name
        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                result = subprocess.run(
                    [lo, "--headless", "--convert-to", "xlsx",
                     input_path, "--outdir", tmp_dir],
                    capture_output=True,
                    timeout=60,
                )
                if result.returncode != 0:
                    stderr_msg = result.stderr.decode(errors='replace').strip()
                    logger.warning(
                        f"[{fname}] LibreOffice 转换失败 (returncode={result.returncode})"
                        f"{': ' + stderr_msg if stderr_msg else ''}"
                    )
                    return None

                # LibreOffice 输出文件名：扩展名会被改为 .xlsx
                # 例如 "AA JE Oct 28.xls" → "AA JE Oct 28.xlsx"
                converted = Path(tmp_dir) / fname
                if not converted.exists():
                    # 尝试匹配 .xlsx 后缀的输出文件
                    candidates = list(Path(tmp_dir).glob("*.xlsx"))
                    if not candidates:
                        all_files = list(Path(tmp_dir).iterdir())
                        logger.warning(
                            f"[{fname}] LibreOffice 转换完成但未找到 .xlsx 输出文件"
                            f"（临时目录内容: {[f.name for f in all_files]}）"
                        )
                        return None
                    converted = candidates[0]

                # 读入内存，避免 tmp_dir 被清理后文件消失
                buf = io.BytesIO(converted.read_bytes())
                buf.seek(0)
                logger.warning(f"[{fname}] 已通过 LibreOffice 重新导出")
                return buf

        except subprocess.TimeoutExpired:
            logger.warning(f"[{fname}] LibreOffice 转换超时（>60s）")
            return None
        except Exception as e:
            logger.warning(f"[{fname}] LibreOffice 转换异常: {type(e).__name__}: {e}")
            return None

    def _try_calamine(
        self,
        input_path: str,
        output_path: str,
        engine: SanitizeEngine,
        stats_before: dict,
    ) -> dict | None:
        """
        用 python-calamine（Rust 引擎）高速读取文件并脱敏

        比 openpyxl read_only 快 20-30 倍，对损坏/格式异常文件容错性强。
        仅保留数据值，不保留样式/透视表。

        返回 stats dict 表示成功，返回 None 表示 calamine 未安装或读取失败。
        """
        try:
            from python_calamine import CalamineWorkbook
        except ImportError:
            return None  # 未安装，静默跳过

        import openpyxl

        fname = Path(input_path).name
        try:
            cal_wb = CalamineWorkbook.from_path(input_path)
        except Exception as e:
            logger.warning(f"[{fname}] calamine 读取失败 → {type(e).__name__}: {e}")
            return None

        logger.warning(f"[{fname}] 使用 calamine 高速读取（样式/透视表不保留）")

        new_wb = openpyxl.Workbook(write_only=True)  # write_only 写入更快
        for sheet_name in cal_wb.sheet_names:
            ws = new_wb.create_sheet(title=sheet_name[:31])
            scale_seed = f"{fname}::{sheet_name}"
            rows = cal_wb.get_sheet_by_name(sheet_name).to_python()
            for row in rows:
                out_row = []
                for val in row:
                    if val is None or val == "":
                        out_row.append(val)
                    elif isinstance(val, bool):
                        out_row.append(val)
                    elif isinstance(val, (datetime, date, time)):
                        out_row.append(val)
                    elif isinstance(val, str):
                        if val.startswith("="):
                            out_row.append(val)
                        elif val.strip():
                            out_row.append(engine.process_text(val))
                        else:
                            out_row.append(val)
                    elif isinstance(val, (int, float)):
                        if math.isnan(val) or math.isinf(val):
                            out_row.append(val)  # NaN/Inf 原样保留
                        elif 2000 <= float(val) <= 2099:
                            out_row.append(val)
                        else:
                            out_row.append(engine.scale_number(val, scale_seed))
                    else:
                        out_row.append(val)
                ws.append(out_row)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        new_wb.save(output_path)
        new_wb.close()
        return self._calc_stats(engine.stats, stats_before)

    def _process_xls(self, input_path: str, output_path: str, engine: SanitizeEngine) -> dict:
        """
        处理 .xls 文件
        使用 xlrd 读取，openpyxl 写出为 .xlsx

        兼容情况：
        - 真正的 BIFF 格式 .xls → xlrd 正常处理
        - 实为 XML 格式但扩展名为 .xls（Excel 2003 XML Spreadsheet）→ 改用 openpyxl 读取
        """
        import xlrd
        import openpyxl

        stats_before = dict(engine.stats)

        try:
            xls_book = xlrd.open_workbook(input_path)
        except xlrd.XLRDError as e:
            # xlrd 无法读取，可能是 Excel 2003 XML Spreadsheet（扩展名 .xls 但实为 XML）
            logger.warning(f"xlrd 无法读取，尝试 LibreOffice 转换: {Path(input_path).name} ({e})")

            # 优先用 LibreOffice 转成标准 xlsx 再处理（速度快，样式完整）
            fname = Path(input_path).name
            lo_buf = self._try_libreoffice_convert(input_path)
            if lo_buf is not None:
                # ── 尝试 1: openpyxl 读取 LibreOffice 转换后的 xlsx（保留样式）──
                try:
                    wb = openpyxl.load_workbook(lo_buf)
                    stats_before_lo = dict(engine.stats)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        scale_seed = f"{Path(input_path).name}::{sheet_name}"
                        try:
                            ws.reset_dimensions()
                        except Exception:
                            pass
                        for row in ws.iter_rows():
                            for cell in row:
                                if cell.value is None:
                                    continue
                                if isinstance(cell.value, (datetime, date, time)):
                                    continue
                                if isinstance(cell.value, str) and cell.value.startswith("="):
                                    continue
                                if isinstance(cell.value, bool):
                                    continue
                                if isinstance(cell.value, str) and cell.value.strip():
                                    cell.value = engine.process_text(cell.value)
                                elif isinstance(cell.value, (int, float)):
                                    fmt = cell.number_format or ""
                                    if self._is_date_format(fmt):
                                        continue
                                    val = float(cell.value)
                                    if 2000 <= val <= 2099:
                                        continue
                                    cell.value = engine.scale_number(cell.value, scale_seed)
                    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
                    # 输出改为 .xlsx 扩展名
                    out = Path(output_path).with_suffix(".xlsx")
                    wb.save(out)
                    wb.close()
                    return self._calc_stats(engine.stats, stats_before_lo)
                except Exception as lo_err:
                    logger.warning(
                        f"[{fname}] LibreOffice 转换后 openpyxl 无法读取 → "
                        f"{type(lo_err).__name__}: {lo_err}"
                    )

                # ── 尝试 2: calamine 读取 LibreOffice 转换后的 xlsx（不保留样式，但快）──
                # lo_buf 是标准 xlsx，calamine 可以读（原始 Excel 2003 XML 它读不了）
                lo_buf.seek(0)
                tmp_lo_path = None
                try:
                    with tempfile.NamedTemporaryFile(
                        suffix=".xlsx", delete=False, prefix="lo_converted_"
                    ) as tmp_f:
                        tmp_f.write(lo_buf.read())
                        tmp_lo_path = tmp_f.name

                    cal_result = self._try_calamine(
                        tmp_lo_path,
                        Path(output_path).with_suffix(".xlsx"),
                        engine,
                        dict(engine.stats),
                    )
                    if cal_result is not None:
                        logger.warning(
                            f"[{fname}] LibreOffice + calamine 联合处理成功（样式不保留）"
                        )
                        return cal_result
                except Exception as e:
                    logger.warning(
                        f"[{fname}] LibreOffice 转换后 calamine 也无法读取 → "
                        f"{type(e).__name__}: {e}"
                    )
                finally:
                    if tmp_lo_path:
                        try:
                            Path(tmp_lo_path).unlink()
                        except OSError:
                            pass

            # calamine 尝试原始文件（对非 Excel 2003 XML 的格式可能有效）
            cal_result = self._try_calamine(
                input_path,
                Path(output_path).with_suffix(".xlsx"),
                engine,
                dict(engine.stats),
            )
            if cal_result is not None:
                return cal_result

            # 最终兜底：内置 ElementTree XML 解析（无需外部工具，但较慢）
            has_lo = self._find_libreoffice() is not None

            if has_lo:
                tip = "（LibreOffice 转换后 openpyxl / calamine 均无法读取，降级到内置解析器）"
            else:
                tip = "\n          提示：安装 LibreOffice 可大幅提速 → brew install --cask libreoffice"

            logger.warning(
                f"[{fname}] 使用内置 XML 解析器（⚠️ 可能较慢，请耐心等待）{tip}"
            )
            return self._process_xls_xml(input_path, output_path, engine)
        xlsx_book = openpyxl.Workbook()

        if xlsx_book.sheetnames:
            xlsx_book.remove(xlsx_book.active)

        for sheet_idx in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)
            scale_seed = f"{Path(input_path).name}::{xls_sheet.name}"

            total_rows = xls_sheet.nrows
            for row_idx in range(total_rows):
                if row_idx > 0 and row_idx % 500 == 0:
                    logger.info(f"[xls] {Path(input_path).name} / {xls_sheet.name}: 已处理 {row_idx}/{total_rows} 行...")
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    cell_type = xls_sheet.cell_type(row_idx, col_idx)

                    if cell_value is None or cell_value == "":
                        continue

                    # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=boolean, 5=error
                    if cell_type == 3:  # 日期 → 跳过
                        # 转换 xlrd 日期为 Python datetime
                        try:
                            date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                            cell_value = datetime(*date_tuple)
                        except Exception:
                            pass
                    elif cell_type == 1:  # 文本
                        text = str(cell_value)
                        if not text.startswith("="):  # 跳过公式
                            cell_value = engine.process_text(text)
                    elif cell_type == 2:  # 数值
                        val = float(cell_value)
                        if not (2000 <= val <= 2099):  # 跳过年份
                            cell_value = engine.scale_number(val, scale_seed)

                    xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        xlsx_book.save(output_path)
        xlsx_book.close()

        return self._calc_stats(engine.stats, stats_before)

    def _process_xls_xml(self, input_path: str, output_path: str, engine: SanitizeEngine) -> dict:
        """
        处理 Excel 2003 XML Spreadsheet 格式
        （文件扩展名为 .xls，实际内容是 XML，xlrd 和 openpyxl 均不支持）

        XML 结构：
        <Workbook>
          <Worksheet ss:Name="Sheet1">
            <Table>
              <Row>
                <Cell><Data ss:Type="String">text</Data></Cell>
                <Cell><Data ss:Type="Number">123</Data></Cell>
              </Row>
            </Table>
          </Worksheet>
        </Workbook>
        """
        import xml.etree.ElementTree as ET
        import openpyxl

        stats_before = dict(engine.stats)

        try:
            tree = ET.parse(input_path)
        except ET.ParseError as e:
            logger.error(f"XML 解析失败: {Path(input_path).name}: {e}")
            return {"entities_found": 0, "entities_replaced": 0}

        root = tree.getroot()

        # Excel 2003 XML 命名空间
        NS = {
            "ss": "urn:schemas-microsoft-com:office:spreadsheet",
            "o":  "urn:schemas-microsoft-com:office:office",
        }
        # 命名空间可能直接在标签里，也可能在属性里，兼容两种写法
        def _tag(ns_uri, local):
            return f"{{{ns_uri}}}{local}"

        xlsx_book = openpyxl.Workbook()
        if xlsx_book.sheetnames:
            xlsx_book.remove(xlsx_book.active)

        # 遍历所有 Worksheet
        for ws_elem in root.iter(_tag("urn:schemas-microsoft-com:office:spreadsheet", "Worksheet")):
            sheet_name = ws_elem.get(
                _tag("urn:schemas-microsoft-com:office:spreadsheet", "Name"), "Sheet"
            )
            xlsx_sheet = xlsx_book.create_sheet(title=sheet_name[:31])  # Excel 限制 31 字符
            scale_seed = f"{Path(input_path).name}::{sheet_name}"

            fname = Path(input_path).name
            row_idx = 0
            for row_elem in ws_elem.iter(_tag("urn:schemas-microsoft-com:office:spreadsheet", "Row")):
                row_idx += 1
                if row_idx % 500 == 0:
                    logger.info(f"[xls-xml] {fname} / {sheet_name}: 已处理 {row_idx} 行...")
                col_idx = 0
                for cell_elem in row_elem.iter(_tag("urn:schemas-microsoft-com:office:spreadsheet", "Cell")):
                    col_idx += 1
                    data_elem = cell_elem.find(
                        _tag("urn:schemas-microsoft-com:office:spreadsheet", "Data")
                    )
                    if data_elem is None or data_elem.text is None:
                        continue

                    cell_type = data_elem.get(
                        _tag("urn:schemas-microsoft-com:office:spreadsheet", "Type"), "String"
                    )
                    raw = data_elem.text.strip()

                    if cell_type == "Number":
                        try:
                            val = float(raw)
                            if not (2000 <= val <= 2099):
                                val = engine.scale_number(val, scale_seed)
                            xlsx_sheet.cell(row=row_idx, column=col_idx, value=val)
                        except ValueError:
                            xlsx_sheet.cell(row=row_idx, column=col_idx, value=raw)
                    elif cell_type == "String":
                        sanitized = engine.process_text(raw)
                        xlsx_sheet.cell(row=row_idx, column=col_idx, value=sanitized)
                    elif cell_type in ("DateTime", "Date"):
                        # 日期直接原样写入
                        xlsx_sheet.cell(row=row_idx, column=col_idx, value=raw)
                    else:
                        xlsx_sheet.cell(row=row_idx, column=col_idx, value=raw)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        xlsx_book.save(output_path)
        xlsx_book.close()

        logger.info(f"Excel 2003 XML 解析完成: {Path(input_path).name}")
        return self._calc_stats(engine.stats, stats_before)

    def _process_csv(
        self,
        input_path: str,
        output_path: str,
        engine: SanitizeEngine,
        ext: str,
    ) -> dict:
        """处理 .csv / .tsv 文件"""
        from sanitizer.utils import detect_encoding

        stats_before = dict(engine.stats)

        encoding = detect_encoding(input_path)
        delimiter = "\t" if ext == ".tsv" else ","
        scale_seed = f"{Path(input_path).name}::data"

        rows = []
        with open(input_path, "r", encoding=encoding, errors="replace", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                rows.append(row)

        sanitized_rows = []
        for row in rows:
            new_row = []
            for cell in row:
                if not cell or not cell.strip():
                    new_row.append(cell)
                    continue

                # 尝试数值缩放
                try:
                    num = float(cell.replace(",", ""))
                    if "." in cell or (len(cell.replace(",", "")) <= 8):
                        # 跳过年份
                        if 2000 <= num <= 2099:
                            new_row.append(cell)
                            continue
                        scaled = engine.scale_number(num, scale_seed)
                        if "," in cell and "." not in cell:
                            new_row.append(f"{int(scaled):,}")
                        elif "." in cell:
                            decimal_places = len(cell.split(".")[-1])
                            new_row.append(f"{scaled:.{decimal_places}f}")
                        else:
                            new_row.append(str(int(scaled)))
                        continue
                except ValueError:
                    pass

                new_row.append(engine.process_text(cell))

            sanitized_rows.append(new_row)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(sanitized_rows)

        return self._calc_stats(engine.stats, stats_before)

    @staticmethod
    def _calc_stats(current: dict, before: dict) -> dict:
        return {
            "entities_found": current["entities_found"] - before["entities_found"],
            "entities_replaced": current["entities_replaced"] - before["entities_replaced"],
        }
